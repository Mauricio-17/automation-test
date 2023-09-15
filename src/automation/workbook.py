from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart.layout import Layout, ManualLayout

book = load_workbook('../data/report_test.xlsx')
work_sheet_primary = book['Primary']

alignment_center = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin'),   # Left border
                right=Side(border_style='thin'),  # Right border
                top=Side(border_style='thin'),    # Top border
                bottom=Side(border_style='thin')) # Bottom borde

font = Font(name='Calibri',
                 size=12,
                 bold=False,
                 italic=False,
                 color='040D12')

font_bold = Font(name='Calibri',
                 size=12,
                 bold=True,
                 italic=False,
                 color='040D12')

fill_pattern_sky = PatternFill(patternType='solid', fgColor='64CCC5')
fill_pattern_green = PatternFill(patternType='solid', fgColor='A6FF96')
fill_pattern_yellow = PatternFill(patternType='solid', fgColor='F8DE22')
fill_pattern_orange = PatternFill(patternType='solid', fgColor='FF9B50')
fill_pattern_red = PatternFill(patternType='solid', fgColor='C01313')
fill_pattern_grey = PatternFill(patternType='solid', fgColor='C4C1A4')

# ------------ Primary worksheet

for row in work_sheet_primary['A1':'J1']:
    for cell in row:
        cell.fill = fill_pattern_sky
        cell.alignment = alignment_center
        cell.border = border
        cell.font = font_bold

for row in work_sheet_primary['A2':'J2']:
    for cell in row:
        cell.fill = fill_pattern_green
        cell.alignment = alignment_center
        cell.border = border
        cell.font = font

for row in work_sheet_primary['A3':'J5']:
    for cell in row:
        cell.fill = fill_pattern_yellow
        cell.alignment = alignment_center
        cell.border = border
        cell.font = font

for row in work_sheet_primary['A6':'J7']:
    for cell in row:
        cell.fill = fill_pattern_orange
        cell.alignment = alignment_center
        cell.border = border
        cell.font = font

for row in work_sheet_primary['A8':'J9']:
    for cell in row:
        cell.fill = fill_pattern_red
        cell.alignment = alignment_center
        cell.border = border
        cell.font = font

for row in work_sheet_primary['A10':'J11']:
    for cell in row:
        cell.fill = fill_pattern_grey
        cell.alignment = alignment_center
        cell.border = border
        cell.font = font

for row in work_sheet_primary['A12':'J12']:
    for cell in row:
        cell.alignment = alignment_center
        cell.border = border
        cell.font = font

for row in work_sheet_primary['A1':'A12']:
    for cell in row:
        cell.font = font_bold

column_dimensions_a = work_sheet_primary.column_dimensions['A']
column_dimensions_b = work_sheet_primary.column_dimensions['B']
column_dimensions_c = work_sheet_primary.column_dimensions['C']
column_dimensions_d = work_sheet_primary.column_dimensions['D']
column_dimensions_e = work_sheet_primary.column_dimensions['E']
column_dimensions_f = work_sheet_primary.column_dimensions['F']
column_dimensions_g = work_sheet_primary.column_dimensions['G']
column_dimensions_h = work_sheet_primary.column_dimensions['H']
column_dimensions_i = work_sheet_primary.column_dimensions['I']
column_dimensions_j = work_sheet_primary.column_dimensions['J']
column_dimensions_a.width = 15.57
column_dimensions_b.width = 15.43
column_dimensions_c.width = 20.14
column_dimensions_d.width = 20.14
column_dimensions_e.width = 20.14
column_dimensions_f.width = 20.14
column_dimensions_g.width = 20.14
column_dimensions_h.width = 20.14
column_dimensions_i.width = 20.14
column_dimensions_j.width = 8.43

# ------------ Secondary worksheet

work_sheet_secondary = book['Secondary']

top_left_cell = work_sheet_secondary["A1"]
top_left_cell.value = "ASESORES"

column_dimensions_sec_a = work_sheet_secondary.column_dimensions['A']
column_dimensions_sec_b = work_sheet_secondary.column_dimensions['B']
column_dimensions_sec_c = work_sheet_secondary.column_dimensions['C']
column_dimensions_sec_d = work_sheet_secondary.column_dimensions['D']
column_dimensions_sec_e = work_sheet_secondary.column_dimensions['E']
column_dimensions_sec_f = work_sheet_secondary.column_dimensions['F']
column_dimensions_sec_g = work_sheet_secondary.column_dimensions['G']

column_dimensions_sec_a.width = 21.30
column_dimensions_sec_b.width = 10.00
column_dimensions_sec_c.width = 14.30
column_dimensions_sec_d.width = 16.14
column_dimensions_sec_e.width = 10.86
column_dimensions_sec_f.width = 8.43
column_dimensions_sec_g.width = 8.43

for row in work_sheet_secondary['A1':'A10']:
    for cell in row:
        cell.font = font_bold
        cell.fill = fill_pattern_sky
        cell.alignment = alignment_center

for row in work_sheet_secondary['B1':'B10']:
    for cell in row:
        cell.font = font
        cell.fill = fill_pattern_green
        cell.alignment = alignment_center
        cell.border = border

for row in work_sheet_secondary['C1':'C10']:
    for cell in row:
        cell.font = font
        cell.fill = fill_pattern_yellow
        cell.alignment = alignment_center
        cell.border = border

for row in work_sheet_secondary['D1':'D10']:
    for cell in row:
        cell.font = font
        cell.fill = fill_pattern_orange
        cell.alignment = alignment_center
        cell.border = border

for row in work_sheet_secondary['E1':'E10']:
    for cell in row:
        cell.font = font
        cell.fill = fill_pattern_red
        cell.alignment = alignment_center
        cell.border = border

for row in work_sheet_secondary['F1':'F10']:
    for cell in row:
        cell.font = font
        cell.fill = fill_pattern_grey
        cell.alignment = alignment_center
        cell.border = border

for row in work_sheet_secondary['G1':'G10']:
    for cell in row:
        cell.font = font
        cell.alignment = alignment_center
        cell.border = border

for row in work_sheet_secondary['B1':'G1']:
    for cell in row:
        cell.font = font_bold

for row in work_sheet_secondary['B10':'G10']:
    for cell in row:
        cell.font = font_bold

book.save("../data/result.xlsx")
book.close()